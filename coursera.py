from lxml import etree
from openpyxl import Workbook
import requests
from bs4 import BeautifulSoup
import json
from os.path import join
from random import sample

COURSES_XML_URL = 'https://www.coursera.org/sitemap~www~courses.xml'
QUANTITY_COURSES_TO_OUTPUT = 20


def convert_soup_to_text(tag):
    return tag.text if tag else None


def get_courses_list():
    xml_page = requests.get(COURSES_XML_URL)
    root = etree.fromstring(xml_page.content)
    links = [link.text for link in root.iter('{*}loc')]
    links = sample(links, QUANTITY_COURSES_TO_OUTPUT)
    return links[:QUANTITY_COURSES_TO_OUTPUT]


def get_average_score_of_course(soup):
    score = convert_soup_to_text(soup.find('div', {'class': 'ratings-text bt3-visible-xs'}))
    return score


def get_datetime_course(soup):
    datetime = None
    json_course = convert_soup_to_text(soup.find('script', {'type': 'application/ld+json'}))
    if json_course:
        datetime = json.loads(json_course)['hasCourseInstance'][0]['startDate']
    return datetime


def get_course_info(course_url):
    page = requests.get(course_url).content
    soup = BeautifulSoup(page,'html.parser')
    course_name = soup.find('div', {'class':'title display-3-text'}).text
    course_lang = soup.find('div', {'class': 'language-info'}).text
    course_date = get_datetime_course(soup)
    duration = len(soup.find_all('div', {'class': 'week'}))
    average_score = get_average_score_of_course(soup)
    return (course_name, course_lang, course_date, duration, average_score)


def output_courses_info_to_xlsx(filepath, courses_info):
    wb = Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1).value = 'Name of course'
    sheet.cell(row=1, column=2).value = 'Language'
    sheet.cell(row=1, column=3).value = 'Date of start'
    sheet.cell(row=1, column=4).value = 'Duration'
    sheet.cell(row=1, column=5).value = 'Average score(out of 5)'
    for number, info_about_course in enumerate(courses_info):
        sheet.cell(row=number+2, column=1).value = info_about_course[0]
        sheet.cell(row=number+2, column=2).value = info_about_course[1]
        sheet.cell(row=number+2, column=3).value = info_about_course[2]
        sheet.cell(row=number+2, column=4).value = info_about_course[3]
        sheet.cell(row=number+2, column=5).value = info_about_course[4]
    wb.save(join(filepath, 'Courses from Coursera.xlsx'))


if __name__ == '__main__':
    links = get_courses_list()
    filepath = input('Enter filepath to directory: \n')
    courses_info = []
    print('Parsing has been started')
    courses_info = [get_course_info(link) for link in links ]
    output_courses_info_to_xlsx(filepath, courses_info)
    print('Done!')